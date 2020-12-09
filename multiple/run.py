import json
from collections import defaultdict

import pymysql
from DBUtils.PooledDB import PooledDB
from openpyxl import Workbook, load_workbook


# CONN_CONFIG = {
# 	"host": "192.168.12.10",
# 	"port": 3306,
# 	"user": "root",
# 	"passwd": "iiecas",
# 	"db": "test",
# 	'charset': 'utf8',
# }

CONN_CONFIG = {
    "host": "192.168.12.9",
    "port": 4002,
    "user": "root",
    "passwd": "eHw@dF2%",
    "db": "test",
    'charset': 'utf8',
}


class SQLPoll(object):

	# docstring for DbConnection

	__poll = None

	def __init__(self):
		self.pool = self.__get_db()
		self.conn = self.pool.connection()
		self.cursor = self.conn.cursor(cursor=pymysql.cursors.DictCursor)

	@classmethod
	def __get_db(cls):
		if cls.__poll is None:
			cls.__poll = PooledDB(
				creator=pymysql,
				mincached=10,
				maxconnections=100,
				**CONN_CONFIG)
		return cls.__poll

	def fetch_all(self, sql, args=None):
		if args is None:
			self.cursor.execute(sql)
		else:
			self.cursor.execute(sql, args)
		result = self.cursor.fetchall()
		return result

	def fetch_one(self, sql, args=None):
		if args is None:
			self.cursor.execute(sql)
		else:
			self.cursor.execute(sql, args)
		result = self.cursor.fetchone()
		return result

	def execute(self, sql, args=None):
		self.cursor.execute(sql, args)
		self.conn.commit()
		result = self.cursor.lastrowid
		return result

	def __close(self):
		self.cursor.close()
		self.conn.close()

	def __enter__(self):
		return self

	def __exit__(self, exc_type, exc_val, exc_tb):
		self.__close()


desc = {
	"简单1": "用户名：win7，密码：qwert",
	"中等2": "被核查的机器为办公用互联网电脑, 用户名：win7，密码：qwert",
	"困难1": "用户名：win7，密码：qwert",
	"困难3": "用户名：root，密码：qwert",
}


def create_data():
	"""pass"""
	path = "核查环境ip规划.xlsx"
	workbook = load_workbook(path)
	sheet = workbook["Sheet1"]
	data = []
	row_list = list(sheet.rows)[1:48]
	for row in row_list:
		row_data = []
		cols = [row[index] for index in (1, 4, 5, 6, 8)]
		for cell in cols:
			row_data.append(cell.value)

		data.append(row_data)

	result = defaultdict(dict)
	keys = ("", "简单1", "中等2", "困难1", "困难3")
	ports = ("", "3389", "3389", "3389", "22")
	for item in data:
		info = item[1:]
		result[item[0]] = {keys[index]: {"entrance": f"{item[index]}:{ports[index]}", "desc": desc[keys[index]]} for index in range(1, len(keys))}

	with open("核查环境ip规划.json", "w") as fd:
		json.dump(result, fd)

	return result


problems_dict = {
	"简单2": "701", "困难3": "700", "中等1": "699", "中等3": "687",
	"中等2": "686", "困难2": "685", "简单1": "684", "困难1": "683"
}

# data = {
# 	"北京": {"样题1": "192.168.100.10:3389"},
# 	"天津": {"样题1": "192.168.101.10:3389"},
# 	"河北": {"样题1": "192.168.102.10:3389"},
# 	"山西": {"样题1": "192.168.100.20:3389"},
# 	"辽宁": {"样题1": "192.168.102.20:3389"},
# 	"吉林": {"样题1": "192.168.100.30:3389"},
# 	"江苏": {"样题1": "192.168.102.30:3389"},
# 	"浙江": {"样题1": "192.168.100.40:3389"},
# 	"安徽": {"样题1": "192.168.101.40:3389"},
# 	"江西": {"样题1": "192.168.102.40:3389"},
# 	"山东": {"样题1": "192.168.100.50:3389"},
# 	"河南": {"样题1": "192.168.101.50:3389"},
# 	"湖北": {"样题1": "192.168.102.50:3389"},
# 	"湖南": {"样题1": "192.168.100.60:3389"},
# 	"广东": {"样题1": "192.168.101.60:3389"},
# 	"广西": {"样题1": "192.168.102.60:3389"},
# 	"海南": {"样题1": "192.168.100.70:3389"},
# 	"四川": {"样题1": "192.168.101.70:3389"},
# 	"贵州": {"样题1": "192.168.102.70:3389"},
# 	"云南": {"样题1": "192.168.100.80:3389"},
# 	"西藏": {"样题1": "192.168.101.80:3389"},
# 	"陕西": {"样题1": "192.168.102.80:3389"},
# 	"甘肃": {"样题1": "192.168.100.90:3389"},
# 	"青海": {"样题1": "192.168.101.90:3389"},
# 	"宁夏": {"样题1": "192.168.102.90:3389"},
# 	"沈阳": {"样题1": "192.168.101.100:3389"},
# 	"大连": {"样题1": "192.168.102.100:3389"},
# 	"长春": {"样题1": "192.168.100.110:3389"},
# 	"上海": {"样题1": "192.168.102.110:3389"},
# 	"南京": {"样题1": "192.168.100.120:3389"},
# 	"重庆": {"样题1": "192.168.101.120:3389"},
# 	"杭州": {"样题1": "192.168.102.120:3389"},
# 	"宁波": {"样题1": "192.168.100.130:3389"},
# 	"厦门": {"样题1": "192.168.101.130:3389"},
# 	"济南": {"样题1": "192.168.102.130:3389"},
# 	"青岛": {"样题1": "192.168.100.140:3389"},
# 	"武汉": {"样题1": "192.168.101.140:3389"},
# 	"广州": {"样题1": "192.168.102.140:3389"},
# 	"深圳": {"样题1": "192.168.100.150:3389"},
# 	"成都": {"样题1": "192.168.101.150:3389"},
# 	"福建": {"样题1": "192.168.102.150:3389"},
# 	"新疆": {"样题1": "192.168.100.160:3389"},
# 	"西安": {"样题1": "192.168.101.160:3389"},
# 	"黑龙江": {"样题1": "192.168.101.30:3389"},
# 	"内蒙古": {"样题1": "192.168.101.20:3389"},
# 	"哈尔滨": {"样题1": "192.168.101.110:3389"},
# 	"新疆兵团": {"样题1": "192.168.100.100:3389"}
# }


# problems_dict = {
# 	"样题1": 718
# }


teams_dict = {
	"广东": "4735", "广州": "4784", "深圳": "4783", "江苏": "4710", "南京": "4970", "湖北": "4729", 
	"武汉": "4971", "江西": "4734", "上海": "4741", "浙江": "4712", "杭州": "4973", "宁波": "4974",
	"湖南": "4730", "安徽": "4728", "广西": "4739", "福建": "4714", "厦门": "4975", "北京": "4702",
	"海南": "4736", "贵州": "4732", "四川": "4731", "成都": "4976", "河北": "4721", "河南": "4725",
	"山东": "4726", "济南": "4977", "青岛": "4747", "云南": "4733", "山西": "4727", "吉林": "4719",
	"长春": "4978", "甘肃": "4722", "辽宁": "4720", "沈阳": "4979", "大连": "4980", "陕西": "4724", 
	"西安": "4981", "宁夏": "4744", "西藏": "4740", "天津": "4742", "重庆": "4743", "新疆": "4737",
	"青海": "4723", "内蒙古": "4738", "黑龙江": "4708", "哈尔滨": "4972", "新疆兵团": "4706"
}


def main():
	"""pass"""
	sql_one = '''
		UPDATE plot_teamproblem 
		SET entrance = %s, entrance_desc = %s
		WHERE
			game_id = 399 
			AND team_id = %s 
			AND origin_id = %s
	'''
	try:
		data = create_data()
		with SQLPoll() as db:
			for name, info in data.items():
				team_id = teams_dict[name]
				for title, item in info.items():
					origin_id = problems_dict[title]
					db.execute(sql_one, (item["entrance"], item["desc"], team_id, origin_id))
	except Exception as e:
		return e

	return True


# def main():
# 	"""pass"""
# 	sql_one = '''
# 		UPDATE plot_teamproblem 
# 		SET entrance = %s, entrance_desc = %s
# 		WHERE
# 			game_id = 413 
# 			AND team_id = %s 
# 			AND origin_id = %s
# 	'''
# 	try:
# 		with SQLPoll() as db:
# 			for name, info in data.items():
# 				team_id = teams_dict[name]
# 				for title, item in info.items():
# 					origin_id = problems_dict[title]
# 					db.execute(sql_one, (item, "用户名：win7，密码：qwert", team_id, origin_id))
# 	except Exception as e:
# 		return e

# 	return True


def create_info():
	"""批量获取分析环境IP地址"""
	path = "核查环境ip规划.xlsx"
	workbook = load_workbook(path)
	sheet = workbook["Sheet1"]
	data = []
	row_list = list(sheet.rows)[1:48]
	for row in row_list:
		row_data = []
		cols = [row[index] for index in (1, 10)]
		for cell in cols:
			row_data.append(cell.value)

		data.append(row_data)

	result = {}
	for item in data:
		result[item[0]] = item[1]

	with open("分析环境ip规划.json", "w") as fd:
		json.dump(result, fd)

	return result


def info_add():
	"""批量发布队伍提示"""
	sql_one = '''
		INSERT INTO a_info ( title, content, pub_time, type, game_id, pub_person, g_p_id, team_id, stage )
		VALUES
			("比赛通知", %s, now(), '管理', 399, 52, 0, %s, 1)
	'''
	data = create_info()
	with SQLPoll() as db:
		for team_name, ip in data.items():
			team_id = teams_dict[team_name]
			content = f"分析环境IP地址： {ip}:3389，用户名：win7，密码：qwert"
			db.execute(sql_one, (content, team_id))

	return True

if __name__ == '__main__':
	# create_info()
	# print(main())
	print(info_add())
